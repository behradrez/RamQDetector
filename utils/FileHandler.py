# Functions to handle XL file
import openpyxl
from utils.PatientBlueprints import  Patient, PatientList


def get_patient_list(File_name):
    NAM_col = ''
    Name_Col = ''
    Actes_Col = ''

    SVisits = 0
    CMVisits = 0

    list_of_patients = []
    list_of_patients_reduced = []

    wb = openpyxl.load_workbook(filename=File_name)
    sheet = wb.active

    # Returns Column number of NAM and Name values
    for x in range(1, 30):
        cell = sheet.cell(row=1, column=x)
        if cell.value is not None:
            if cell.value == "NAM":
                NAM_col = x - 1
            elif cell.value == "Patient":
                Name_Col = x - 1
            elif cell.value == "Actes" or "Actes" in cell.value:
                Actes_Col = x - 1
    for row in sheet:
        Patient_NAM = row[NAM_col].value
        Patient_name = row[Name_Col].value
        Patient_Actes = row[Actes_Col].value

        if Patient_NAM == "NAM" and Patient_name == "Patient" and Patient_Actes == "Actes":
            continue

        elif Patient_name is None and Patient_NAM is None and Patient_Actes is None:
            print("Progress: 20%")
            break

        if Patient_NAM is not None:
            Patient_NAM = Patient_NAM.upper()

        if Patient_name is not None:
            Patient_name = Patient_name.upper()

        if Patient_Actes is not None:
            if 'S' in Patient_Actes:
                SVisits += Patient_Actes.count('S')
            elif 'C' in Patient_Actes or 'M' in Patient_Actes:
                CMVisits += (Patient_Actes.count('C') + Patient_Actes.count('M'))

        newPatient = Patient(Patient_name, Patient_NAM, 1)
        list_of_patients.append(newPatient)

    for x in range(len(list_of_patients)):
        Visits = 1
        for y in range(len(list_of_patients)):
            if list_of_patients[x] == list_of_patients[y] and x != y:
                Visits += 1
        list_of_patients[x].numVisits = Visits
        if list_of_patients[x] not in list_of_patients_reduced:
            list_of_patients_reduced.append(list_of_patients[x])
    return PatientList(list_of_patients_reduced, (SVisits, CMVisits, len(list_of_patients)))


def XLFile_Creator(PatientList_OBJ, final_Name):
    wb = openpyxl.Workbook()
    ws = wb.active

    PatientList_OBJ.Sort_List()
    list_Of_Patients = PatientList_OBJ.listOfPatients

    ws['B1'].value = "NAM"
    ws['C1'].value = "Patient"
    ws["D1"].value = "# Visites"
    ws['F2'].value = "S visits: " + str(PatientList_OBJ.S_Visits)
    ws['F3'].value = "C/M visits: " + str(PatientList_OBJ.CM_Visits)
    ws['F4'].value = "Total visits: " + str(PatientList_OBJ.total_Visits)

    for i in range(2, len(list_Of_Patients) + 2):
        ws['B' + str(i)].value = list_Of_Patients[i - 2].NAM
        ws['C' + str(i)].value = list_Of_Patients[i - 2].name
        ws['D' + str(i)].value = list_Of_Patients[i - 2].numVisits
    wb.save(str(final_Name) + ".xlsx")


def FileHandler(path_to_file: str, fileName: str) -> None:
    XLFile_Creator(get_patient_list(path_to_file), fileName)
